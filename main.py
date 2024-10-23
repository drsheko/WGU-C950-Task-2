#  Student Name : Shady Israel
#  Student ID: 012069822

import openpyxl
from openpyxl import Workbook, load_workbook
from symtable import Class
import datetime
import csv

print('Please wait, Program is loading ............. 2%')

# Hash Table to store packages data, each package data is connected to unique key (package ID)
class HashTable:
    def __init__(self):
        self.table = {}
    # Insert method that takes  two parameters packageID and package class
    # Key => packageID , Value => WGUPackage class including Address,city,zip, deadline, Weight and delivery status
    def insert(self, package_id, package):
        self.table[package_id] = package
    # Lookup method that takes packageID as a parameter and return package data
    def lookup(self, package_id):
        return self.table.get(package_id)

# Initialize hash table
hash_table = HashTable()

# Class to build each package that takes ID, Street, city, state, zip, weight, deadline as parameters
# with default values of status, Note, departure time, delivery time to be assigned later
class WGUPackages:
    def __init__(self, pID, pAddress, pCity, pState, pZip, pWeight, pDeadline, pStatus= "at the hub", pNote = None, pDeparture= None, pDelivery =None ):
        self.pID = pID
        self.pAddress = pAddress
        self.pCity = pCity
        self.pState = pState
        self.pZip = pZip
        self.pWeight = pWeight
        self.pDeadline = pDeadline
        self.pStatus = pStatus
        self.pNote = pNote
        self.pDeparture = None
        self.pDelivery = None


# Function to get the distance between two addresses to be used to get the closest package location from the current location of truck
def getDistance(address1, address2):
    # Reading distance info from WGUPS distance table
    file1 = load_workbook('WGUPS Distance Table.xlsx')
    distanceInfoFile = file1.active
    rows = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
    foundRow = 35
    foundCol = 'AC'
    for i in range(26):
        cellIndex = 'A' + str(i+9)
        if (address1 in distanceInfoFile[cellIndex].value):
            foundRow = i+9
            for j in range(27):
                cellIndex = rows[j] + str(8)
                if (address2 in distanceInfoFile[cellIndex].value):
                    foundCol = rows[j]
                    i=40
                    break
    distanceIndex = foundCol + str(foundRow)
    foundDistance = distanceInfoFile[distanceIndex].value
    # If the cell is empty rerun the function after switching addresses
    if (foundDistance is None ):
        return getDistance(address2, address1)
    else:
        return int(foundDistance)


# This creates a package using the data from the Provided Package file and push each package to the hash table
def getDataFromPackage():
    # Reading packages info from WGUPS Package file
    file1 = load_workbook('WGUPS Package File.xlsx')
    packagesInfoFile = file1.active
    rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for i in range(40):
        packageInfo =[]
        for j in range(7):
            cellIndex = rows[j] + str(i + 9)
            packageInfo.append(packagesInfoFile[cellIndex].value)

        #Set All package properities
        ID = int(packageInfo[0])
        address = packageInfo[1]
        city = packageInfo[2]
        state = packageInfo[3]
        postal = packageInfo[4]
        weight = packageInfo[5]
        deadline = packageInfo[6]
        # Insert the package into hashtable associated with its unique packageID key
        hash_table.insert(ID, WGUPackages(ID, address, city, state, postal, weight, deadline))

# Call getDataFromPackage Function
getDataFromPackage()

# Class to build WGUPS TRUCKS, each truck has its miles, its current location, departure time, packages
class Trucks:
    def __init__(self,  Miles, Location, Departure, Packages):
        self.Miles = Miles
        self.Location = Location
        self.Time = Departure
        self.Departure = Departure
        self.Packages = Packages

# Variable to store loading status
wait = 8

# Function to deliver truck package by checking the closest package location to be delivered next, update package info in the hash table
def deliverTruckPackages(truck):
    # Array to store truck packages
    packagesToBeDelivered = []
    # For Loop to get each package data from Hash table into packagesToBeDelivered array based on PackageID key
    for packageID in truck.Packages:
        package = hash_table.lookup(packageID)
        packagesToBeDelivered.append(package)

    # Loop to deliver all  packages in packagesToBeDelivered list, Will break when there is no package left to deliver
    while len(packagesToBeDelivered) > 0:
        global wait
        wait =wait + 2
        print('Please wait, Program is loading ............. ' + str(wait) + '%')
        # Set nextAddress to a random high number
        nextAddress = 5000
        # Initiate value of next package is none
        nextPackage = None

        # Loop to determine which package should be delivered next
        for thePackage in packagesToBeDelivered:
            # Exception 2 packages with id 6, 25 because they will arrive to hub late and need to be delivered before 10:30 am
            # these two packages should be delivered first
            if thePackage.pID in [25,6]:
                nextPackage = thePackage
                nextAddress = getDistance(truck.Location, thePackage.pAddress)
                break
            #Update nextAddress based on the closest package address to the current location of truck
            if getDistance(truck.Location, thePackage.pAddress) <= nextAddress:
                nextAddress = getDistance(truck.Location, thePackage.pAddress)
                nextPackage = thePackage

        # Remove delivered package from packagesToBeDelivered array
        packagesToBeDelivered.remove(nextPackage)
        # Add miles to deliver the current package to total truck miles
        truck.Miles += nextAddress
        #Update truck location
        truck.Location = nextPackage.pAddress
        # Update Truck time by adding required time to deliver the packgae based on truck's speed (18 mph) to current truck time
        truck.Time += datetime.timedelta(hours=nextAddress / 18)
        #Update package delivery time to be stored in Hash table later
        nextPackage.pDelivery = truck.Time
        # Update package departure time to be stored in Hash table later
        nextPackage.pDeparture = truck.Departure
        # Save Updated Package Info in HASH TABLE
        hash_table.table[nextPackage.pID] = nextPackage

    # Add Distance and time required for truck to return from last deliverd package location to the hub
    nextAddress = getDistance(truck.Location, "4001 South 700 East")
    truck.Miles += nextAddress
    truck.Time += datetime.timedelta(hours=nextAddress / 18)

# Define Truck Class Instance for first and second Truck with packages, starting time and starting location
firstTruck = Trucks ( 0.0, "4001 South 700 East", datetime.timedelta (hours=8), [1,2, 13, 14, 15, 16, 19, 20, 27, 29, 30, 31, 34, 37, 40])
secondTruck = Trucks( 0.0, "4001 South 700 East", datetime.timedelta(hours=8), [3, 4, 5,18, 26,28, 35,36,38])
thirdTruck = Trucks ( 0.0, "4001 South 700 East", datetime.timedelta (hours=10, minutes=5), [6,7,8,9,10,11,12,17, 21,22,23,24,25, 32, 33,39])

# Call deliverTruckPackages on the first truck
deliverTruckPackages (firstTruck)
# Call deliverTruckPackages on the second truck
deliverTruckPackages (secondTruck)

# Get next available driver after coming back to hub to drive the third truck
availableDriver = min(firstTruck.Time, secondTruck.Time)
# Set third trick departure time equal to the next available driver at the hub to be used as a departure time of third truck
thirdTruck.Departure = availableDriver
thirdTruck.Time = availableDriver
# Call deliverTruckPackages on the Third truck
deliverTruckPackages (thirdTruck)

# Print Class Title, Student Name and ID
print("\n \n WGUPS Project \n C950 Task 2 \n Student Name : Shady Israel \n Student ID: 012069822 ")
print('\n ====================================================================================\n')

# Print total miles for ALL TRUCKS
print("The total miles for all trucks are: ", (firstTruck.Miles + secondTruck.Miles+thirdTruck.Miles))
print('\n ====================================================================================\n')

#Function to Get Package Status based on packageID and Time, The returned value will be used in displayData Function
def getPackageStatus(packageID, certainTime):
    deliveryTime = str(hash_table.table[packageID].pDelivery)
    # Convert delivery time string into array [hours, minuts]
    myTime = deliveryTime.split(':')
    hour = myTime[0]
    minute = myTime[1]
    # Set Delivery time to be compared with the provided certain time later
    delivery = datetime.timedelta(hours=int(hour), minutes=int(minute))

    departureTime = str(hash_table.table[packageID].pDeparture)
    # Convert departure time string into array [hours, minuts]
    myTime =  departureTime.split(':')
    hour = myTime[0]
    minute = myTime[1]
    # Set departure time to be compared with the provided certain time later
    departure = datetime.timedelta(hours=int(hour), minutes=int(minute))

    #Processing Certain time to be compared wit both delivery and departure time
    (hour, minute) = certainTime.split(':')
    certainTime = datetime.timedelta(hours=int(hour), minutes=int(minute))

    #IF Package Delivery time is before the provided time that means the package has been delivered already
    if(certainTime > delivery):
        return 'Package has been delivered!'
    # IF Package Delivery time is after the provided time and Package Departure time is before the provided time
    # That means the package is out for delivery but has not been delivered yet
    elif (certainTime < delivery and certainTime > departure):
        return 'Package is en route!'
    # If Not delivered nor en route => the package is still at the hub
    else:
        return 'Package is at the hub!'



# Function to display Package Data at certain time
def displayData(packageID, certainTime):
    if(certainTime ==''):
        print("\n Error!! Your input is not valid, Please Enter Valid Time \n")
        return
    # If No PackageId provided, display Data for ALL packages
    if(packageID == ''):
        for i in range (1,41):
            status = getPackageStatus(i, certainTime)
            print('ID:'+ str(hash_table.table[i].pID)+
                  '\n======= \nDelivery Address:'
                  + hash_table.table[i].pAddress
                  + hash_table.table[i].pCity
                  + hash_table.table[i].pState
                  + str(hash_table.table[i].pZip)
                  + ', Deadline:' + str(hash_table.table[i].pDeadline)
                  + ', STATUS:' + status
                  + ', Departure:' + str(hash_table.table[i].pDeparture)
                  + ', Delivery Time:' + str(hash_table.table[i].pDelivery)+'\n'

                  )
    else:
        # If provided PackageID out of range => Display Error Message with First package Data and Recall the function
        i = int(packageID)
        if i > 40 :
            print('Sorry, We don`t have package with such ID, we will display the package with ID = 1 ')
            i=1
        status = getPackageStatus(i, certainTime)

        print('ID:'+ str(hash_table.table[i].pID)+
                  '\n ======= \nDelivery Address:'
                  + hash_table.table[i].pAddress
                  + hash_table.table[i].pCity
                  + hash_table.table[i].pState
                  + str(hash_table.table[i].pZip)
                  + ', Deadline:' + str(hash_table.table[i].pDeadline)
                  + ', STATUS:' + status
                  + ', Departure:' + str(hash_table.table[i].pDeparture)
                  + ', Delivery Time:' + str(hash_table.table[i].pDelivery) +'\n'
              )

# Loop To collect Input for Package Id and time
while True:
    # print(first Truck.tMiles + secondTruck.tMiles third Truck.tMiles))
    packageIdInput = input("To check the status for ALL Packages, Please press Enter: "
                       "\nFor the status for only One Package, Please Enter Package ID then hit Enter:")

    timeInput = input("Please Enter time to check the status of selected/all package(s) in format HH:MM :\n")

    # Call displayData Function with User Input of PackageID and TIME
    displayData(packageIdInput, timeInput)




